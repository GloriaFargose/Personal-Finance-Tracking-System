import tkinter
from tkinter import ttk,messagebox,simpledialog
from tkcalendar import DateEntry
import os
import openpyxl
import matplotlib.pyplot as plt


root=tkinter.Tk()
root.title("Personal Finance Management System")

title =tkinter.Label(root, font=(" Times new roman", 18, "bold italic"),text="Personal Finance Management System")
title.pack()


frame=tkinter.Frame(root)
frame.pack()

expenses=[]

user_info_frame=tkinter.LabelFrame(frame,text="User Information")
user_info_frame.grid(row=0, column=0, padx=20, pady=20)

first_name_label=tkinter.Label(user_info_frame, text="First Name:")
first_name_label.grid(row=0,column=1)
last_name_label=tkinter.Label(user_info_frame, text="Last Name:")
last_name_label.grid(row=0,column=2)

first_name_entry=tkinter.Entry(user_info_frame)
last_name_entry=tkinter.Entry(user_info_frame)
first_name_entry.grid(row=1, column=1)
last_name_entry.grid(row=1, column=2)

title_label=tkinter.Label(user_info_frame, text="Title")
title_combobox=ttk.Combobox(user_info_frame,values=["Mr.","Ms.","Mrs.","Prefer not to mention"])
title_label.grid(row=0, column=0)                           
title_combobox.grid(row=1,column=0)

salary_label=tkinter.Label(user_info_frame, text="Monthly Salary:")
salary_label.grid(row=3, column=0)

salary_entry=tkinter.Entry(user_info_frame)
salary_entry.grid(row=4, column=0)

for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)
    
###Expense
    
expense_frame=tkinter.LabelFrame(frame,text="Expenditure Information")
expense_frame.grid(row=1,column=0, sticky="news",padx=30,pady=30)

currency_label=tkinter.Label(expense_frame, text="Select Currency:")
currency_combobox=ttk.Combobox(expense_frame,values=["USD","INR","EUR","CAD"])
currency_label.grid(row=1, column=0)                           
currency_combobox.grid(row=2,column=0)


your_expense_label=tkinter.Label(expense_frame, text="Expense Amount:")
your_expense_label.grid(row=1,column=1)
your_expense_entry=tkinter.Entry(expense_frame)
your_expense_entry.grid(row=2, column=1)

item_description_label=tkinter.Label(expense_frame, text="Item Description:")
item_description_label.grid(row=1,column=2)
item_description_entry=tkinter.Entry(expense_frame)
item_description_entry.grid(row=2, column=2)

category_label=tkinter.Label(expense_frame, text="Category:")
category_combobox=ttk.Combobox(expense_frame,values=["Rent","Utilities","Travel","Food","Groceries","Online Subscriptions","Other"])
category_label.grid(row=3, column=0)                           
category_combobox.grid(row=4,column=0)

total_label=tkinter.Label(expense_frame, text="Total Expenses:")
total_label.grid(row=9,column=0)

total_savings_label=tkinter.Label(expense_frame, text="Total Savings:")
total_savings_label.grid(row=9,column=2)


def pick_date():
    selected_date= cal.get()
    date_label.config(text=selected_date)


cal=DateEntry(expense_frame,setmode="day",date_pattern="mm/dd/yyyy")
cal.grid(row=4,column=1,padx=15)


date_button=tkinter.Button(expense_frame,text="Select",command=pick_date)
date_button.grid(row= 5, column=1)

date_label=tkinter.Label(expense_frame,text="Date:")
date_label.grid(row=3,column=1)


for widget in expense_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# Scroll Bar
frame_list=tkinter.Frame(root)
frame_list.pack()

scrollbar=tkinter.Scrollbar(expense_frame)
scrollbar.grid(row=6,column=2,sticky= "ns")

expense_listbox=tkinter.Listbox(expense_frame, yscrollcommand=scrollbar.set,width=40,height=10)
expense_listbox.grid(row=6,column=1,padx=30, pady=10)

scrollbar.config(command=expense_listbox.yview)



###Functions


def add_expense():
    currency= currency_combobox.get()
    expense= your_expense_entry.get()
    description= item_description_entry.get()
    category= category_combobox.get()
    date= cal.get()
    if expense and date:
        expense_listbox.insert(tkinter.END, f"{currency} - {expense} - {description} - {category} - {date}")
        expenses.append((float(expense),currency, description, category, date))
        your_expense_entry.delete(0,tkinter.END)
        item_description_entry.delete(0,tkinter.END)
        cal.delete(0,tkinter.END)
        update_total()

    else:
        messagebox.showwarning("WARNING", "Expense and Date cannot be empty.")
    


def edit_expense():
    selected_index=expense_listbox.curselection()
    if selected_index:
        selected_index = selected_index[0]
        selected_expense=expenses[selected_index]
        if len(selected_expense)>= 4:
            new_expense=simpledialog.askstring("Edit Expense","Enter new expense:", initialvalue=selected_expense[0])
            new_category=simpledialog.askstring("Edit Category","Enter new category:",initialvalue=selected_expense[3])
            if new_expense:
                expenses[selected_index] = (
                     new_expense,
                     selected_expense[1],
                     selected_expense[2],
                     new_category,
                     selected_expense[4],
                     )
                refresh_list()
                update_total()
            else:
                messagebox.showwarning("Warning","Selected expense does not have enough elements.")
            
            
def delete_expenses():
    selected_index=expense_listbox.curselection()
    if selected_index:
        selected_index = selected_index[0]
        del expenses[selected_index]
        expense_listbox.delete(selected_index)
        update_total()
        
def refresh_list():
    expense_listbox.delete(0,tkinter.END)
    for expense,currency, description, category,date in expenses:
        expense_listbox.insert(tkinter.END, f"{currency_combobox.get()} - {expense} - {description}-{category} - {date}")

    
def update_total():
    total_expenses = sum(float(expense[0]) for expense in expenses)
    total_label.config(text=f"Total Expenses:" + currency_combobox.get()+ " " + f"{total_expenses:.2f}")

savings=0       
def total_savings():
    if salary_entry.get() and expenses:
        montly_salary=float(salary_entry.get())
        total_expenses=sum(float(expense[0]) for expense in expenses)
        savings= montly_salary - total_expenses
        total_savings_label.config(text=f"Total Savings:{currency_combobox.get()} {savings:.2f}")

        
filepath = r"/Users/gloriafargose/Library/CloudStorage/OneDrive-DePaulUniversity/IS411/MyExpenses.xlsx"

if not os.path.exists(filepath):
    workbook=openpyxl.Workbook()
    sheet=workbook.active
    heading=["Title","First Name","Last Name","Monthly Salary", "Currency","Expense Amount","Item Description","Category","Date","Total Savings"]
    sheet.append(heading)
    workbook.save(filepath)

def save_expenses():
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.active
    while(sheet.max_row > 1):
        sheet.delete_rows(2)
    for i in range(expense_listbox.size()):
        item = expense_listbox.get(i)
        split_item=item.split("-")
        sheet.append([title_combobox.get(), first_name_entry.get(), last_name_entry.get(), salary_entry.get(),
                  split_item[0], split_item[1], split_item[2],
                  split_item[3], split_item[4], total_savings_label.cget('text')])

    workbook.save(filepath)


def show_expenses():
    category={}
    for expense in expenses:
        if expense[3] in category:
            category[expense[3]]+=float(expense[0])
        else:
            category[expense[3]]= float(expense[0])

    labels=category.keys()
    values=category.values()
    plt.figure(figsize=(6,6))
    plt.pie(values,labels=labels,autopct="%1.1f%%")
    plt.title("Monthly Expense Distribution")

    plt.show()
    
#Buttons
add_button=tkinter.Button(expense_frame,text="Enter Expense",command= add_expense)
add_button.grid(row= 4, column=2)

edit_button=tkinter.Button(expense_frame,text="Edit Expense",command= edit_expense)
edit_button.grid(row=8, column=2)

delete_button=tkinter.Button(expense_frame,text="Delete Expenses",command=delete_expenses)
delete_button.grid(row=8, column=0)

save_button=tkinter.Button(expense_frame,text="Save Expenses",command=save_expenses)
save_button.grid(row=11, column=1)

savings_button=tkinter.Button(expense_frame,text="Calculate Total Savings",command=total_savings)
savings_button.grid(row=9, column=1)

chart_button=tkinter.Button(expense_frame,text="Show Expenses",command=show_expenses)
chart_button.grid(row=13, column=1)

    
root.mainloop()
